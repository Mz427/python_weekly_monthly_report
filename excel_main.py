from openpyxl import Workbook
from openpyxl import load_workbook
from print_date_of_week import DateOfWeek

'''
wb_week = Workbook()
ws = wb_week.active

ws['A1'] = 42
ws.append([1, 2, 3])
d = ws.cell(row=4, column=2, value=10)
d = ws.cell(row=3, column=2).value = "这是第三行, 第二列"
ws.append([1, 2, 3])
cell_range = ws['A1' : 'C2']
col_range = ws['C' : 'D']
for row in cell_range:
	for cell in row:
		print(cell.value)

for row in ws.values:
	for cell in row:
		print(cell)

for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
	print(row)

#tuple(ws.rows)
#tuple(ws.columns)

wb_week.save("sample.xlsx")
'''

#wb_week = Workbook()

mz_today = DateOfWeek()

for i in mz_today.get_week_date():
    print(i)

for i in mz_today.sheet_name_date():
    print(i)
